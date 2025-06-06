from collections import defaultdict
import csv
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
from tabulate import tabulate
from datetime import datetime, timedelta
import numpy as np 
import matplotlib.pyplot as plt 
import sqlite3
from sqlite3 import Error
import sys 
from datetime import datetime
from tabulate import tabulate
import csv
import pandas as pd

db_nombre = "Negocio_Rentas_Bicicletas.db"
lista_colores = ["rojo", "violeta", "azul", "verde", "amarillo", "naranja"]
prestamos = {}

def calcular_moda(arr):
    valores, conteos = np.unique(arr, return_counts=True)
    indice = np.argmax(conteos)
    return valores[indice]

def verificacion_fechas():
    fecha_inicial = input("Ingrese la fecha inicial (mm-dd-yyyy): ")
    fecha_final = input("Ingrese la fecha final (mm-dd-yyyy): ")
    fecha_inicial_dt = datetime.strptime(fecha_inicial, "%m-%d-%Y")
    fecha_final_dt = datetime.strptime(fecha_final, "%m-%d-%Y")
    return fecha_inicial_dt, fecha_final_dt, True

def parse_fecha_prestamo(fecha_str):
    formatos = ['%m-%d-%Y', '%m/%d/%Y']
    for formato in formatos:
        try:
            return datetime.strptime(fecha_str, formato)
        except ValueError:
            continue
    raise ValueError(f"No fue posible parsear la fecha: {fecha_str}")

def tabla_existe(cursor, nombre_tabla):
    cursor.execute("SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?", (nombre_tabla,))
    return cursor.fetchone() is not None

def exportar_csv(lista_datos, headers, opcion):
    
    if opcion in ["prestamos", "unidades", "clientes"]:
        nombre_archivo = f"{opcion}.csv"
    else:
        fecha = datetime.now().strftime("%Y%m%d%H%M%S")
        nombre_archivo = f"{opcion}_{fecha}.csv"
    
   
    with open(nombre_archivo, 'w', encoding='latin1', newline='') as archivo:
        writer = csv.writer(archivo)
       
        writer.writerow(headers)
        
        for row in lista_datos:
            writer.writerow(row)
    print("Archivo exportado con éxito.")

def exportar_excel(lista_datos, headers, opcion):
    
    fecha = datetime.now().strftime("%Y%m%d%H%M%S")
    nombre_archivo = f"{opcion}_{fecha}.xlsx"
    
    book = openpyxl.Workbook()
    sheet = book["Sheet"]
    sheet.title = "Hoja 1"
   
    negritas = Font(bold=True)
    centro = Alignment(horizontal='center', vertical='center')
    borde = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = negritas
        cell.alignment = centro
        cell.border = borde
   
    for row_idx, row_data in enumerate(lista_datos, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = centro
            cell.border = borde
    
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width
    
    for col in range(1, len(headers) + 1):
        sheet.cell(row=1, column=col).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thick'))
   
    book.save(nombre_archivo)
    print("Archivo exportado con éxito.")

def exportar_csv_desde_db(opcion):
    if opcion in ["prestamos", "unidades", "clientes"]:
        nombre_archivo = f"{opcion}.csv"
        nombre_tabla = opcion.upper()
    else:
        fecha = datetime.now().strftime("%Y%m%d%H%M%S")
        nombre_archivo = f"{opcion}_{fecha}.csv"
        nombre_tabla = opcion

    try:
        con = sqlite3.connect(db_nombre)
        cursor = con.cursor()

        if not tabla_existe(cursor, nombre_tabla):
            print(f"La tabla {nombre_tabla} no existe en la base de datos.")
            return

        cursor.execute(f"PRAGMA table_info({nombre_tabla})")
        headers = [columna[1] for columna in cursor.fetchall()]

        cursor.execute(f"SELECT * FROM {nombre_tabla}")
        rows = cursor.fetchall()

        with open(nombre_archivo, 'w', encoding='latin1', newline='') as archivo:
            writer = csv.writer(archivo)
            writer.writerow(headers)
            writer.writerows(rows)

        print(f"Archivo {nombre_archivo} exportado con éxito.")

    except sqlite3.Error as e:
        print(f"Error al acceder a la base de datos: {e}")
    except Exception as e:
        print(f"Error al exportar el archivo: {e}")
    finally:
        con.close()

def exportar_excel_desde_db(opcion):
    fecha = datetime.now().strftime("%Y%m%d%H%M%S")
    nombre_archivo = f"{opcion}_{fecha}.xlsx"
    nombre_tabla = opcion.upper()

    try:
        con = sqlite3.connect(db_nombre)
        cursor = con.cursor()

        if not tabla_existe(cursor, nombre_tabla):
            print(f"La tabla {nombre_tabla} no existe en la base de datos.")
            return

        cursor.execute(f"PRAGMA table_info({nombre_tabla})")
        headers = [columna[1] for columna in cursor.fetchall()]

        cursor.execute(f"SELECT * FROM {nombre_tabla}")
        rows = cursor.fetchall()

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Hoja 1"

        negritas = Font(bold=True)
        centro = Alignment(horizontal='center', vertical='center')
        borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = negritas
            cell.alignment = centro
            cell.border = borde

        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = centro
                cell.border = borde

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

        book.save(nombre_archivo)
        print(f"Archivo {nombre_archivo} exportado con éxito.")

    except sqlite3.Error as e:
        print(f"Error al acceder a la base de datos: {e}")
    except Exception as e:
        print(f"Error al exportar el archivo: {e}")
    finally:
        con.close()

def exportar_bd(opcion):
    while True:
        exportar_input = input("¿Desea exportar? [S/n]: ")
        if exportar_input not in ["S", "s", "N", "n"]:
            print("Opción incorrecta, ingresar de nuevo.")
            continue
        elif exportar_input.strip().lower() == "s":
            print("\nOpciones: ")
            print("1. CSV")
            print("2. Excel")
            opcion_exportar = input("Seleccione una opción: ")
            if opcion_exportar == "1":
                exportar_csv_desde_db(opcion)
                break
            elif opcion_exportar == "2":
                exportar_excel_desde_db(opcion)
                break
            else:
                print("Opción inválida, ingresar de nuevo.")
        else:
            break

def exportar_default(datos, headers, opcion):
    while True:
        exportar = input("¿Desea exportar? [S/n]: ")
        if exportar not in ["S", "s", "N", "n"]:
            print("Opción incorrecta, ingresar de nuevo.")
            continue
        elif exportar.strip().lower() == "s":
            print("\nOpciones: ")
            print("1. CSV")
            print("2. Excel")
            opcion_exportar = input("Seleccione una opción: ")
            if opcion_exportar == "1":
                exportar_csv(datos, headers, opcion)
                break
            elif opcion_exportar == "2":
                exportar_excel(datos, headers, opcion)
                break
            else:
                print("Opción inválida, ingresar de nuevo.")
        else:
            break

def exportar_datos_csv_colores(color, unidades_con_color):
    csv_filename = f'unidades_{color}.csv'
    try:
        with open(csv_filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(['Unidad', 'Rodada'])
            writer.writerows(unidades_con_color)
        print(f'Datos exportados exitosamente a {csv_filename}.')
    except Exception as e:
        print(f"Error al exportar a CSV: {e}")

def exportar_datos_excel_colores(color, unidades_con_color):
    excel_filename = f'unidades_{color}.xlsx'
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(['Unidad', 'Rodada'])
        for unidad, rodada in unidades_con_color:
            ws.append([unidad, rodada])
        wb.save(excel_filename)
        print(f'Datos exportados exitosamente a {excel_filename}.')
    except Exception as e:
        print(f"Error al exportar a Excel: {e}")

def registrar_cliente():
    validacion_caracteres = r'^[A-Za-z\s]+$'
    validacion_longitud = r'^.{1,40}$'
    validacion_telefono = r'^\d{10}$'

    while True:
        apellidos = input("Ingresar apellidos (hasta 40 caracteres), teclear X para cancelar: ")
        if apellidos.upper() == "X":
            return
        if re.match(validacion_longitud, apellidos):
            if re.match(validacion_caracteres, apellidos):
                break
            print("Error: Los apellidos no pueden contener números ni caracteres especiales.")
            continue
        print("Error: Los apellidos son obligatorios y deben tener hasta 40 caracteres.")

    while True:
        nombres = input("Ingrese nombre (hasta 40 caracteres), teclear X para cancelar: ")
        if nombres.upper() == "X":
            return
        if re.match(validacion_longitud, nombres):
            if re.match(validacion_caracteres, nombres):
                break
            print("Error: Los nombres no pueden contener números ni caracteres especiales.")
            continue
        print("Error: Los nombres son obligatorios y deben tener hasta 40 caracteres.")
    
    while True:
        telefono = input("Ingrese su número teléfonico (10 dígitos), teclear X para cancelar: ")
        if telefono.upper() == "X":
            return
        if re.match(validacion_telefono, telefono):
            break
        print("Error: El teléfono debe contener exactamente 10 dígitos numéricos.")
    
    f_apellidos = apellidos.strip().upper()
    f_nombres = nombres.strip().upper()

    try:
        with sqlite3.connect("Negocio_Rentas_Bicicletas.db") as con:
            cursor = con.cursor()
            valores = (f_apellidos, f_nombres, telefono)
            cursor.execute("INSERT INTO CLIENTE (apellidos, nombre, telefono) \
        VALUES(?, ?, ?)", valores)
            print(f"La clave asignada fue {cursor.lastrowid}")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        con.close()

def registrar_unidad():
    validacion_rodada = r'^2[069]$'

    while True:
        rodada = input("Ingrese rodada [20, 26, 29], teclear X para cancelar: ")
        if rodada.upper() == "X":
            return
        if re.match(validacion_rodada, rodada):
            break
        else:
            print("Error: La rodada debe ser 20, 26 o 29.")
    while True:
        color = input(f'Ingrese un color de las siguientes opciones [rojo, violeta, azul, verde, amarillo, naranja], teclear X para cancelar: ')
        if color.upper() == "X":
            return
        if color.lower() in lista_colores:
            f_color = color.upper()
            break
        else:
            print('Color inválido. Ingresar nuevamente.')

    try:
        with sqlite3.connect("Negocio_Rentas_Bicicletas.db") as con:
            cursor = con.cursor()
            valores = (rodada, f_color)
            cursor.execute("INSERT INTO UNIDAD (rodada, color) \
        VALUES(?, ?)", valores)
            print(f"La clave asignada fue {cursor.lastrowid}")
    except Error as e:
        print (e)
    except Exception:
        print(f"Se produjo el siguiente error: {sys.exc_info()[0]}")
    finally:
        con.close()

def registrar_prestamo():
    try:
        with sqlite3.connect("Negocio_Rentas_Bicicletas.db",
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as con:
            cursor = con.cursor()
            cursor.execute("""
                SELECT p.folio, p.claveUnidad, p.fechaPrestamo, p.diasPrestados 
                FROM PRESTAMO p 
                WHERE p.estado = 'ACTIVO'
            """)
            prestamos_activos = cursor.fetchall()

            cursor.execute("SELECT claveUnidad, rodada, color FROM UNIDAD")
            unidades_data = cursor.fetchall()
            unidades_headers = ["Clave", "Rodada", "Color"]

            cursor.execute("SELECT claveCliente, nombre || ' ' || apellidos as nombre_completo, telefono FROM CLIENTE")
            clientes_data = cursor.fetchall()
            clientes_headers = ["Clave", "Nombre", "Número"]
            
            while True:
                if unidades_data and clientes_data:
                    print("Unidades: ")
                    print(tabulate(unidades_data, headers=unidades_headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
                else:
                    if not unidades_data:
                        print("No hay unidades registradas por el momento.")
                    else:
                        print("No hay clientes registrados por el momento.")
                    return
                
                try:
                    clave_unidad = input("Ingrese la clave de la unidad, teclear X para cancelar: ")
                    if clave_unidad.upper() == "X":
                        return
                    else:
                        clave_unidad = int(clave_unidad)
                    cursor.execute("SELECT 1 FROM UNIDAD WHERE claveUnidad = ?", (clave_unidad,))
                    if cursor.fetchone():
                        break
                    else:
                        print("Error: La clave ingresada no corresponde a una unidad registrada.")
                except ValueError:
                    print("Error: Debe ingresar un valor numérico.")

            while True:
                print("\nClientes:")
                print(tabulate(clientes_data, headers=clientes_headers, tablefmt='grid', colalign=("left",), disable_numparse=True))

                try:
                    clave_cliente = input("Ingrese la clave del cliente, teclear X para cancelar: ")
                    if clave_cliente.upper() == "X":
                        return
                    else:
                        clave_cliente = int(clave_cliente)
                    cursor.execute("SELECT 1 FROM CLIENTE WHERE claveCliente = ?", (clave_cliente,))
                    if cursor.fetchone():
                        break
                    else:
                        print("Error: No existe ningún cliente con esa clave.")
                except ValueError:
                    print("Error: Debe ingresar un valor numérico.")

            while True:
                fecha_prestamo = input("Ingrese la fecha del préstamo (MM-DD-AAAA), deje en blanco para registrar la fecha actual o teclear X para cancelar: ")
                if fecha_prestamo.upper() == "X":
                    return
                if fecha_prestamo == "":
                    fecha_prestamo = datetime.now().strftime("%m-%d-%Y")
                    fecha_prestamo_dt = datetime.strptime(fecha_prestamo, "%m-%d-%Y")
                    break
                try:
                    if "/" in fecha_prestamo:
                        partes = fecha_prestamo.split("/")
                        fecha_prestamo = f"{partes[0]}-{partes[1]}-{partes[2]}"
                    fecha_prestamo_dt = datetime.strptime(fecha_prestamo, "%m-%d-%Y")
                    if fecha_prestamo_dt.date() >= datetime.now().date():
                        break
                    else:
                        print("Error: La fecha del préstamo no puede ser anterior a la fecha actual.")
                except ValueError:
                    print("Error: Fecha no válida. Debe estar en formato mm-dd-aaaa o mm/dd/aaaa.")

            while True:
                try:
                    dias_prestamo = input("Ingrese la cantidad de días del préstamo (1 a 14), teclear X para cancelar: ")
                    if dias_prestamo.upper() == "X":
                        return
                    else:
                        dias_prestamo = int(dias_prestamo)
                    if 1 <= dias_prestamo <= 14:
                        break
                    else:
                        print("Error: La cantidad de días debe ser entre 1 y 14.")
                except ValueError:
                    print("Error: Debe ingresar un valor numérico.")

            fecha_retorno_dt = fecha_prestamo_dt + timedelta(days=dias_prestamo)
            fecha_retorno = fecha_retorno_dt.strftime('%m-%d-%Y')
            salida = True
            for prestamo in prestamos_activos:
                if clave_unidad == prestamo[1]:
                    fecha_prestamo_activa = prestamo[2].date()
                    fecha_retorno_activa = (prestamo[2] + timedelta(days = prestamo[3])).date()
                    if fecha_retorno_dt.date() <= fecha_prestamo_activa or fecha_prestamo_dt.date() >= fecha_retorno_activa:
                        continue
                    else:
                        salida = False
                        print("Esta unidad ya está reservada en el periodo indicado, no procede registro.")
                        break

            if salida:
                cursor.execute("""
                    INSERT INTO PRESTAMO (claveUnidad, claveCliente, fechaPrestamo, diasPrestados, 
                                        fechaRetorno, estado)
                    VALUES (?, ?, ?, ?, NULL, 'ACTIVO')
                """, (clave_unidad, clave_cliente, fecha_prestamo_dt, dias_prestamo))
                
                con.commit()
                print("Préstamo registrado con éxito.")
                print(f"Fecha de inicio del prestamo: {fecha_prestamo}")
                print(f"Fecha de compromiso de retorno: {fecha_retorno}")
    except sqlite3.Error as e:
        print(f"Error con la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")

def registrar_retorno():
    try:
        with sqlite3.connect("Negocio_Rentas_Bicicletas.db",
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as con:
            cursor = con.cursor()

            while True:
              
                cursor.execute("""
                    SELECT folio, claveCliente, claveUnidad
                    FROM PRESTAMO
                    WHERE estado = 'ACTIVO' AND fechaPrestamo <= ?
                """, (datetime.now(),))
                folios_activos = cursor.fetchall()

                
                tabla_datos = []
                for folio, clave_cliente, clave_unidad in folios_activos:
                    cursor.execute("SELECT apellidos, nombre FROM CLIENTE WHERE claveCliente = ?", (clave_cliente,))
                    cliente = cursor.fetchone()
                    if cliente:
                        nombre_cliente = f"{cliente[0]} {cliente[1]}"
                    else: 
                        nombre_cliente = "Desconocido"
                    tabla_datos.append([folio, nombre_cliente, clave_unidad])

                headers = ["Folio", "Nombre", "Unidad"]

                if tabla_datos:
                    print(tabulate(tabla_datos, headers=headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
                else:
                    print("No hay préstamos por retornar.")
                    break

               
                try:
                    folio = input("Ingrese el folio del préstamo a retornar, teclear X para cancelar: ")
                    if folio.upper() == "X":
                        return
                    folio = int(folio)
                except ValueError:
                    break

                
                cursor.execute("SELECT * FROM PRESTAMO WHERE folio = ?", (folio,))
                prestamo = cursor.fetchone()

                if prestamo:
                    
                    if prestamo[-1] == "ACTIVO":  
                        fecha_retorno = datetime.now().strftime("%m-%d-%Y")
                        fecha_retorno_dt = datetime.strptime(fecha_retorno, "%m-%d-%Y")
                        
                        cursor.execute("UPDATE PRESTAMO SET estado = ?, fechaRetorno = ? WHERE folio = ?", ("RETORNADO", fecha_retorno_dt, folio))

                        con.commit()
                        print("Unidad retornada con éxito.")
                        break
                    else:
                        print("Error: Este préstamo ya fue retornado.")
                else:
                    print("Error: No se encontró un préstamo con ese folio.")
    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")

def reporte_clientes():
    headers = ["Clave", "Apellido(s)", "Nombre", "Teléfono"]
    print("Reporte clientes:")
    try:
        con = sqlite3.connect(db_nombre)
        cursor = con.cursor()
        cursor.execute("SELECT claveCliente, apellidos, nombre, telefono FROM CLIENTE")
        clientes = cursor.fetchall()
        if not clientes:
            print("No hay registros de clientes por el momento.")
        else:
            tabla = [[int(cliente[0]), cliente[1], cliente[2], cliente[3]] for cliente in clientes]
            print("\n" + tabulate(tabla, headers=headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
            exportar_bd("CLIENTE")
    except sqlite3.Error as e:
        print(f"Error de base de datos: {e}")
    finally:
        con.close()

def reporte_clientes_especifico():
    try:
        with sqlite3.connect(db_nombre,
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as con:
            cursor = con.cursor()
            cursor.execute("SELECT claveCliente, apellidos, nombre, telefono FROM CLIENTE")
            clientes = cursor.fetchall()
            if not clientes:
                print("No hay registros de clientes por el momento.")
            else:
                while True:
                    headers = ["Clave", "Apellido(s)", "Nombre", "Teléfono"]
                    tabla = [[int(cliente[0]), cliente[1], cliente[2], cliente[3]] for cliente in clientes]
                    print("Clientes:")
                    print("\n" + tabulate(tabla, headers=headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
                    try:
                        clave_cliente = input("Ingrese la clave del cliente, teclear X para cancelar: ")
                        if clave_cliente.upper() == "X":
                            return
                        else:
                            clave_cliente = int(clave_cliente)
                        cursor.execute("SELECT 1 FROM CLIENTE WHERE claveCliente = ?", (clave_cliente,))
                        if cursor.fetchone():
                            break
                        else:
                            print("Error: No existe ningún cliente con esa clave.")
                    except ValueError:
                        print("Error: Debe ingresar un valor numérico.")
                cursor.execute("SELECT nombre || ' ' || apellidos as nombre_completo, telefono FROM CLIENTE where claveCliente = ?", (clave_cliente,))
                nombre_completo, telefono = cursor.fetchone()
                datos_cliente = f'Nombre: {nombre_completo} | Número de teléfono: {telefono}'
                cursor.execute("""
                    SELECT folio, claveUnidad, fechaPrestamo, diasPrestados, fechaRetorno, estado
                    FROM PRESTAMO
                    WHERE claveCliente = ?
                """, (clave_cliente,))
                historial_cliente = cursor.fetchall()
                tabla_datos = []
                for folio, clave_unidad, fecha_prestamo, dias_prestados, fecha_retorno, estado in historial_cliente:
                    tabla_datos.append([folio, clave_unidad, fecha_prestamo, dias_prestados, fecha_retorno, estado])
                
                headers = ["Folio", "Clave de unidad", "Fecha de préstamo", "Días prestados", "Fecha de retorno", "Estado"]

                if tabla_datos:
                    print(datos_cliente)
                    print(tabulate(tabla_datos, headers=headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
                    exportar_default(tabla_datos, headers, nombre_completo)
                else:
                    print("No hay préstamos registrados para el cliente indicado.")
                

    except sqlite3.Error as e:
        print(f"Error con la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")

def reporte_unidades():
    headers = ["Folio", "Rodada", "Color"]
    tabla_datos = []

    try:
        con = sqlite3.connect(db_nombre)
        cursor = con.cursor()
        cursor.execute("SELECT claveUnidad, rodada, color FROM UNIDAD")
        unidades = cursor.fetchall()
        for unidad in unidades:
            tabla_datos.append([unidad[0], unidad[1], unidad[2]])
        tabla = tabulate(tabla_datos, headers=headers, tablefmt="grid", colalign=("left",), disable_numparse=True)
        print(tabla)
  
        exportar = input("¿Desea exportar los datos? (csv/excel/no): ").lower()
        while True:
            if exportar == 'csv':
                exportar_csv_desde_db('UNIDAD')  
                break
            elif exportar == 'excel':
                exportar_excel_desde_db('UNIDAD')  
                break
            elif exportar == 'no':
                print("No se exportaron los datos.")
                break
            else:
                print("Opción no válida. Intente de nuevo.")
                exportar = input("¿Desea exportar los datos? (csv/excel/no): ").lower()

    except sqlite3.Error as e:
        print(f"Error al acceder a la base de datos: {e}")
    finally:
        if con:
            con.close()

def reporte_por_rodada():
    headers = ["Folio", "Rodada", "Color"]
    tabla_datos = []
    try:
        con = sqlite3.connect(db_nombre)
        cursor = con.cursor()

        while True:
            rodada = input("Ingrese la rodada que desea consultar [20, 26, 29]: ")
            if rodada in ['20', '26', '29']:
                rodada = int(rodada)
                break
            else:
                print("Rodada inválida, intente de nuevo.")
        
        cursor.execute("SELECT claveUnidad, rodada, color FROM UNIDAD WHERE rodada=?", (rodada,))
        unidades_filtradas = cursor.fetchall()
        
        if not unidades_filtradas:
            print(f"No hay unidades registradas para la rodada {rodada}.")
            return
        
        for indice, unidad in enumerate(unidades_filtradas, start=1):
            tabla_datos.append([indice, unidad[1], unidad[2]])  
        
        print("\n" + tabulate(tabla_datos, headers=headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
        
        while True:
            exportar = input("¿Desea exportar los datos filtrados? (csv/excel/no): ").lower()
            if exportar == 'csv':  
                try:
                    con = sqlite3.connect(db_nombre)
                    cursor = con.cursor()
                    cursor.execute(f"PRAGMA table_info(UNIDAD)")
                    headers = [columna[1] for columna in cursor.fetchall()]
                    cursor.execute(f"SELECT claveUnidad, rodada, color FROM UNIDAD WHERE rodada=?", (rodada,))
                    rows = cursor.fetchall()
                    with open(f'UNIDAD_POR_RODADA{rodada}.csv', 'w', encoding='latin1', newline='') as archivo:
                        writer = csv.writer(archivo)
                        writer.writerow(headers)
                        writer.writerows(rows)
                    print(f"Archivo 'UNIDAD' exportado con éxito.")

                except sqlite3.Error as e:
                    print(f"Error al acceder a la base de datos: {e}")
                except Exception as e:
                    print(f"Error al exportar el archivo: {e}")
                finally:
                    if con:
                        con.close()  
                break

            elif exportar == 'excel':    
                try:
                    con = sqlite3.connect(db_nombre)
                    cursor = con.cursor()

                    cursor.execute(f"PRAGMA table_info(UNIDAD)")
                    headers = [columna[1] for columna in cursor.fetchall()]

                    cursor.execute(f"SELECT claveUnidad, rodada, color FROM UNIDAD WHERE rodada = ?", (rodada,))
                    rows = cursor.fetchall()

                    book = openpyxl.Workbook()
                    sheet = book.active
                    sheet.title = "Hoja 1"

                    negritas = Font(bold=True)
                    centro = Alignment(horizontal='center', vertical='center')
                    borde = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                    for col, header in enumerate(headers, start=1):
                        cell = sheet.cell(row=1, column=col)
                        cell.value = header
                        cell.font = negritas
                        cell.alignment = centro
                        cell.border = borde

                    for row_idx, row_data in enumerate(rows, start=2):
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                            cell.alignment = centro
                            cell.border = borde

                    for column in sheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        adjusted_width = (max_length + 2) * 1.2
                        sheet.column_dimensions[column_letter].width = adjusted_width

                    book.save(f'UNIDAD_POR_RODADA{rodada}.xlsx')
                    print(f"Archivo UNIDAD_POR_RODADA{rodada} exportado con éxito.")

                except sqlite3.Error as e:
                    print(f"Error al acceder a la base de datos: {e}")
                except Exception as e:
                    print(f"Error al exportar el archivo: {e}")
                finally:
                    if con:
                        con.close()
                break
            elif exportar == 'no':
                print("No se exportaron los datos.")
                break
            else:
                print("Opción no válida. Por favor, elija 'csv', 'excel' o 'no'.")

    except sqlite3.Error as e:
        print(f"Error al acceder a la base de datos: {e}")
    finally:
        if con:
            con.close()

def reporte_color():
    try:
        with sqlite3.connect(db_nombre) as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT DISTINCT color FROM UNIDAD;")
            lista_colores = [row[0] for row in mi_cursor.fetchall()]
            print(f'\nListado de colores: {lista_colores}')

            while True:
                color = input('Ingrese el color que desea visualizar: ').upper()

                if color in lista_colores:  
                    
                    mi_cursor.execute("""
                        SELECT claveUnidad, rodada 
                        FROM UNIDAD 
                        WHERE color = ?;
                    """, (color,))
                    unidades_con_color = mi_cursor.fetchall()  

                    if unidades_con_color:
                        
                        print(f'\nUnidades con el color "{color}":')
                        headers = ["Unidad", "Rodada"]
                        print(tabulate(unidades_con_color, headers=headers, tablefmt="grid", colalign=("left",), disable_numparse=True))

                        
                        exportar = input('¿Desea exportar los resultados a CSV, Excel o ambos? (csv/excel/ambos/n): ').lower()

                        if exportar == 'csv':
                            exportar_datos_csv_colores(color, unidades_con_color)  
                        elif exportar == 'excel':
                            exportar_datos_excel_colores(color, unidades_con_color)  
                        elif exportar == 'ambos':
                            exportar_datos_csv_colores(color, unidades_con_color)  
                            exportar_datos_excel_colores(color, unidades_con_color)  
                        else:
                            print('No se realizará ninguna exportación.')

                        break  

                    else:
                        print('El color no se encuentra registrado en ninguna unidad.')
                        return

                else:
                    print('El color ingresado no es válido.')
                    continue

    except Exception as e:
        print(f"Se produjo el siguiente error: {e}")

def reporte_retrasos(nombre_bd):
    retrasos = {}
    hoy = datetime.now()
    tabla = []  

    
    try:
        conn = sqlite3.connect(nombre_bd,
                               detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        cursor = conn.cursor()

        
        cursor.execute("""
            SELECT p.folio, p.claveUnidad, p.claveCliente, p.fechaPrestamo, p.diasPrestados,
                   p.fechaRetorno, u.rodada, u.color, c.nombre, c.apellidos, c.telefono
            FROM PRESTAMO p
            JOIN UNIDAD u ON p.claveUnidad = u.claveUnidad
            JOIN CLIENTE c ON p.claveCliente = c.claveCliente
        """)

        prestamos = cursor.fetchall()

        for prestamo in prestamos:
            folio = prestamo[0]
            clave_cliente = prestamo[2]
            fecha_prestamo = prestamo[3]
            dias_prestados = prestamo[4]
            fecha_retorno = prestamo[5]

            

            if fecha_retorno:
                fecha_acordada_retorno = fecha_prestamo + timedelta(days = dias_prestados) 
                if fecha_acordada_retorno < fecha_retorno:
                    dias_retraso = (fecha_retorno - fecha_acordada_retorno).days
                
                    retrasos[folio] = {
                        'unidad': prestamo[6],  
                        'color': prestamo[7],
                        'nombre_completo': f"{prestamo[8]} {prestamo[9]}",  
                        'telefono': prestamo[10],
                        'fecha_prestamo': fecha_prestamo.strftime("%m/%d/%Y"),
                        'fecha_acordada_retorno': fecha_acordada_retorno.strftime("%m/%d/%Y"),
                        'fecha_retorno': fecha_retorno.strftime("%m/%d/%Y"),
                        'dias_retraso': dias_retraso
                    }

    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    finally:
        if conn:
            conn.close()

    
    if retrasos:
        print("Reporte de préstamos con retraso:")
        
        
        for folio, datos in retrasos.items():
            fila = [
                folio,
                datos['unidad'],
                datos['color'],
                datos['nombre_completo'],
                datos['telefono'],
                datos['fecha_prestamo'],
                datos['fecha_acordada_retorno'],
                datos['fecha_retorno'],
                datos['dias_retraso']
            ]
            tabla.append(fila)

        
        tabla.sort(key=lambda x: x[-1], reverse=True)
        
        
        encabezados = [
            "Folio", "Unidad", "Color", "Cliente", "Teléfono", "Fecha de préstamo", "Fecha acordada de retorno", "Fecha de retorno", "Días de retraso"
        ]
        
        
        print(tabulate(tabla, headers=encabezados, tablefmt="grid", colalign=("left",), disable_numparse=True))
    else:
        print("No hay préstamos con retraso.")

    while True:
        opcion = input(f'\n¿Desea exportar? (csv/excel/no): ')

        if opcion.lower() == 'csv':
            if tabla:  
                exportar_csv(tabla, encabezados, 'Reporte_Retrasos')
                print('Datos exportados exitosamente.')
            else:
                print("No hay datos para exportar.")
            break
        elif opcion.lower() == 'excel':
            if tabla:  
                exportar_excel(tabla, encabezados, 'Reporte_Retrasos')
                print('Datos exportados exitosamente.')
            else:
                print("No hay datos para exportar.")
            break
        elif opcion.lower() == 'no': 
            print('Datos no exportados')
            break
        else:
            print('Opción inválida.')

    return retrasos

def reporte_prestamos_por_retornar():

    headers = ["folio", "clave_unidad", "clave_cliente", "fecha_prestamo", "dias_prestamo"]

    try: 
        with sqlite3.connect("Negocio_Rentas_Bicicletas.db",
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT folio, claveUnidad, claveCliente, fechaPrestamo, diasPrestados
                FROM PRESTAMO 
                WHERE estado = 'ACTIVO'
            """)            

            prestamos_raw = cursor.fetchall()
        prestamos_por_retornar = {}

        try:
            fecha_inicial_dt, fecha_final_dt, status = verificacion_fechas()
        except Exception as e:
            print("Error al verificar las fechas:", e)
            return

        print("Reporte de préstamos por retornar:")

        for prestamo in prestamos_raw:
            fecha_prestamo_dt = prestamo[3]

            if fecha_inicial_dt <= fecha_prestamo_dt <= fecha_final_dt:
                folio = prestamo[0]
                prestamos_por_retornar[folio] = prestamo 

        if not prestamos_por_retornar:
            print("No hay préstamos por retornar en el período especificado.")
            return
                
        tabla = [[prestamo[0], prestamo[1], prestamo[2], prestamo[3].strftime("%m/%d/%Y"), prestamo[4]] 
                for prestamo in prestamos_por_retornar.values()]
                
        print("\n" + tabulate(tabla, headers=headers, tablefmt='grid', disable_numparse=True))

        exportar_default(tabla, headers, "reporte_por_retornar")

    except sqlite3.Error as e:
        print(f"Error con la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")
        
def reporte_prestamos_por_periodo():
    headers = ["Folio", "Clave unidad", "Rodada", "Fecha prestamo", "Nombre", "Telefono"]
    prestamos = dict()
    try:     
        with sqlite3.connect(db_nombre,
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT p.folio, p.claveUnidad, u.rodada, p.fechaPrestamo, c.nombre || ' ' || c.apellidos, c.telefono
                FROM PRESTAMO p
                JOIN UNIDAD u ON p.claveUnidad = u.claveUnidad 
                JOIN CLIENTE c on p.claveCliente = c.claveCliente
            """)            

            prestamos_raw = cursor.fetchall()
        
        while True:
            try:
                fecha_inicial_dt, fecha_final_dt, status = verificacion_fechas()
                break 
            except Exception as e:
                print("Error al verificar las fechas:", e)
                return

        print(f'Reporte de préstamos en el periodo del {fecha_inicial_dt.strftime("%m-%d-%Y")} al {fecha_final_dt.strftime("%m-%d-%Y")}:')
        
        for prestamo in prestamos_raw:
            folio = prestamo[0]
            fecha_prestamo_dt = prestamo[3]            
            
            if fecha_inicial_dt <= fecha_prestamo_dt <= fecha_final_dt:
                prestamos[folio] = prestamo 
        
        if not prestamos:
            print("No hay préstamos en el periodo especificado.")
            return


        tabla = [[prestamo[0], prestamo[1], prestamo[2],
                 prestamo[3], prestamo[4], prestamo[5]] 
                for prestamo in prestamos.values()]

        print("\n" + tabulate(tabla, headers=headers, tablefmt='grid', colalign=("left",), disable_numparse=True))
    
        exportar_default(tabla, headers, "rep_periodo")

    except sqlite3.Error as e:
        print(f"Error con la base de datos: {e}")
    except Exception as e:
        print(f"Error: {e}")

def duracion_prestamo():
    resultados = []
    errores = []
    duraciones = []
    
    try:
        conn = sqlite3.connect(db_nombre,
                               detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        cursor = conn.cursor()

        
        cursor.execute("SELECT folio, fechaPrestamo, diasPrestados FROM PRESTAMO")
        prestamos = cursor.fetchall()

        for folio, fecha_prestamo_dt, dias_prestados in prestamos:
            try:
                fecha_retorno = fecha_prestamo_dt + timedelta(days=int(dias_prestados))
                diferencia_dias = (fecha_retorno - fecha_prestamo_dt).days
                resultados.append([folio, fecha_prestamo_dt.strftime('%Y-%m-%d'), fecha_retorno.strftime('%Y-%m-%d'), diferencia_dias])
                duraciones.append(diferencia_dias)
            except ValueError as e:
                errores.append([folio, str(e)])

    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    finally:
        if conn:
            conn.close()
    
    if resultados:
        print("\nResumen de préstamos:")
        print(tabulate(resultados, headers=["Folio", "Fecha Préstamo", "Fecha Retorno", "Duración (días)"], tablefmt="grid"))

    if errores:
        print("\nErrores de parseo:")
        print(tabulate(errores, headers=["Folio", "Error"], tablefmt="grid"))

    if duraciones:
        headers = ['Medida', 'Valor']
        duraciones_np = np.array(duraciones)
        estadisticas = [
            ["Media", np.mean(duraciones_np)],
            ["Mediana", np.median(duraciones_np)],
            ["Moda", calcular_moda(duraciones_np)],
            ["Minimo", np.min(duraciones_np)],
            ["Maximo", np.max(duraciones_np)],
            ["Desviacion Estándar", np.std(duraciones_np)],
            ["Primer Cuartil (Q1)", np.percentile(duraciones_np, 25)],
            ["Tercer Cuartil (Q3)", np.percentile(duraciones_np, 75)]
        ]
        print("\nEstadísticas de duración de préstamos (en días):")
        print(tabulate(estadisticas, headers=["Medida", "Valor"], tablefmt="grid"))

        exportar_default(estadisticas, headers, "Analisis_Duracion")

def ranking_clientes():

    prestamos_por_cliente = {}

    try:
        conn = sqlite3.connect(db_nombre)
        cursor = conn.cursor()

        
        cursor.execute("SELECT claveCliente FROM PRESTAMO")
        prestamos = cursor.fetchall()

       
        for (clave_cliente,) in prestamos:
            prestamos_por_cliente[clave_cliente] = prestamos_por_cliente.get(clave_cliente, 0) + 1

        
        clientes_ranking = []
        cursor.execute("SELECT claveCliente, nombre, apellidos, telefono FROM CLIENTE")
        clientes = cursor.fetchall()

        for clave_cliente, nombre, apellidos, telefono in clientes:
            num_prestamos = prestamos_por_cliente.get(clave_cliente, 0)
            nombre_completo = f"{nombre} {apellidos}"
            clientes_ranking.append([num_prestamos, clave_cliente, nombre_completo, telefono])

       
        clientes_ranking.sort(reverse=True, key=lambda x: x[0])

        
        print("Ranking de Clientes (Ordenado por cantidad de préstamos):")
        headers = ["Núm. de Rentas", "Clave Cliente", "Nombre Completo", "Teléfono"]
        print(tabulate(clientes_ranking, headers=headers, tablefmt="grid"))

    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    finally:
        if conn:
            conn.close()

def reporte_prestamos_por_rodada():
    db_name = "Negocio_Rentas_Bicicletas.db"
    prestamos_por_rodada = {}

    try:
        conn = sqlite3.connect(db_name)
        cursor = conn.cursor()

       
        cursor.execute("SELECT claveUnidad FROM PRESTAMO")
        prestamos = cursor.fetchall()

        if prestamos: 

          
            for (clave_unidad,) in prestamos:
                cursor.execute("SELECT rodada FROM UNIDAD WHERE claveUnidad = ?", (clave_unidad,))
                rodada = cursor.fetchone()
                if rodada:
                    rodada = rodada[0]
                    prestamos_por_rodada[rodada] = prestamos_por_rodada.get(rodada, 0) + 1

            
            prestamos_rodada_list = [[rodada, cantidad] for rodada, cantidad in prestamos_por_rodada.items()]
            prestamos_rodada_list.sort(reverse=True, key=lambda x: x[1])

          
            headers = ["Rodada", "Cantidad de Préstamos"]
            print("Reporte de Préstamos por Rodada:")
            print(tabulate(prestamos_rodada_list, headers=headers, tablefmt="grid"))

           
            rodadas = [item[0] for item in prestamos_rodada_list]
            cantidades = [item[1] for item in prestamos_rodada_list]

            plt.figure(figsize=(8, 6))
            plt.pie(cantidades, labels=rodadas, autopct='%1.1f%%', startangle=140)
            plt.title("Distribución de préstamos por rodada")
            plt.axis('equal')  
            plt.show()

        else:
            print('\nNo hay prestamos registrados\n')
            return    

    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    finally:
        if conn:
            conn.close()

def analisis_color():
    db_name = "Negocio_Rentas_Bicicletas.db"
    prestamos_por_color = {}

   
    try:
        conn = sqlite3.connect(db_name)
        cursor = conn.cursor()

        
        cursor.execute("SELECT claveUnidad FROM PRESTAMO")
        prestamos = cursor.fetchall()

        if prestamos: 

           
            for (clave_unidad,) in prestamos:
                cursor.execute("SELECT color FROM UNIDAD WHERE claveUnidad = ?", (clave_unidad,))
                color = cursor.fetchone()
                if color:
                    color = color[0]
                    prestamos_por_color[color] = prestamos_por_color.get(color, 0) + 1

           
            prestamos_color_list = [[color, cantidad] for color, cantidad in prestamos_por_color.items()]
            prestamos_color_list.sort(reverse=True, key=lambda x: x[1])

           
            headers = ["Color", "Cantidad de Préstamos"]
            print("Reporte de Préstamos por Color:")
            print(tabulate(prestamos_color_list, headers=headers, tablefmt="grid"))

          
            colores = [item[0] for item in prestamos_color_list]
            cantidades = [item[1] for item in prestamos_color_list]

            plt.figure(figsize=(8, 6))
            plt.pie(cantidades, labels=colores, autopct='%1.1f%%', startangle=140)
            plt.title("Distribución de Préstamos por Color")
            plt.axis('equal')  
            plt.show()

        else:
            print('\nNo hay prestamos registrados\n')
            return    


    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    finally:
        if conn:
            conn.close()

def dia_semana():

    prestamos_por_dia = defaultdict(int)

    try:
        conn = sqlite3.connect("Negocio_Rentas_Bicicletas.db", 
                             detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        cursor = conn.cursor()


        cursor.execute("SELECT fechaPrestamo FROM PRESTAMO")
        prestamos = cursor.fetchall()

        if prestamos: 


            for (fecha_str,) in prestamos:
                
                try:
                    if isinstance(fecha_str, str):
                        fecha = datetime.strptime(fecha_str, '%m-%d-%Y')
                    else:
                        fecha = fecha_str
                    
                    
                    dia_semana = fecha.weekday()
                    
                    dia_semana = (dia_semana + 1) % 7
                    prestamos_por_dia[dia_semana] += 1
                except Exception as e:
                    print(f"Error procesando fecha {fecha_str}: {e}")

            
            dias_semana_espanol = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
            prestamos_acumulados = [prestamos_por_dia[i] for i in range(7)]

            
            print("Datos crudos de préstamos por día:", dict(prestamos_por_dia))

            
            if sum(prestamos_acumulados) == 0:
                print("No se encontraron préstamos en la base de datos.")
                return

            
            headers = ["Día de la Semana", "Cantidad de Préstamos"]
            dias_report = list(zip(dias_semana_espanol, prestamos_acumulados))
            print("\nReporte de Préstamos por Día de la Semana:")
            print(tabulate(dias_report, headers=headers, tablefmt="grid"))

            
            plt.figure(figsize=(10, 6))
            plt.bar(dias_semana_espanol, prestamos_acumulados, color='skyblue')
            plt.title("Cantidad de Préstamos por Día de la Semana")
            plt.xlabel("Día de la Semana")
            plt.ylabel("Cantidad de Préstamos")
            plt.xticks(rotation=45)
            plt.grid(axis='y', linestyle='--')
            plt.tight_layout()
            plt.show()

        else:
            print('\nNo hay prestamos registrados\n')
            return    

    except sqlite3.Error as e:
        print(f"Error de SQLite: {e}")
    except Exception as e:
        print(f"Error general: {e}")
    finally:
        if 'conn' in locals() and conn:
            conn.close()

def menu():
    while True:
        print('\nRuta: Menú principal')
        print("\nMENÚ PRINCIPAL")
        print("1. Registro")
        print("2. Préstamo")
        print("3. Retorno")
        print("4. Informes")
        print("5. Salir")

        opcion = input("Seleccione una opción: ")

        if opcion == "1":
            while True:
                print('\nRuta: Menú principal -> Registro')
                print("\nRegistro")
                print("1. Unidad")
                print("2. Cliente")
                print("3. Regresar al menú")
                sub_opcion = input("Seleccione una opción: ")

                if sub_opcion == "1":
                    print('\nRuta: Menú principal -> Registro -> Unidad')
                    registrar_unidad()
                elif sub_opcion == "2":
                    print('\nRuta: Menú principal -> Registro -> Cliente')
                    registrar_cliente()
                elif sub_opcion == "3":
                    break
                else:
                    print("Opción no válida.")
                
        elif opcion == "2":
            if not registrar_unidad:
                print("No hay unidades registradas.")
                    
            if not registrar_cliente:
                    print("No hay clientes registrados.")
            else:
                while True:
                    print('\nRuta: Menú principal -> Préstamo')
                    print("\nPréstamo")
                    print("1. Registrar préstamo")
                    print("2. Regresar al menú")
                    sub_opcion = input("Seleccione una opcion: ")

                    if sub_opcion =="1":
                        print('\nRuta: Menú principal -> Préstamo -> Registrar préstamo')
                        registrar_prestamo()
                    elif sub_opcion =="2":
                        break
                    else:
                        print("Opción no valida")

        elif opcion == "3":
            try:
                print('\nRuta: Menú principal -> Retorno')
                registrar_retorno()
            except IndexError:
                print('No hay prestamos por retornar')

        elif opcion == "4":
            while True:
                print('\nRuta: Menú principal -> Informes')
                print('\nInformes')
                print('1. Reportes')
                print('2. Análisis')
                print('3. Volver al menú principal')
                sub_opcion = input('Seleccione una opción: ')
                if sub_opcion == '1':
                    while True:
                        print('\nRuta: Menú principal -> Informes -> Reportes')
                        print("\nReportes")
                        print("1. Clientes")
                        print('2. Listado de unidades')
                        print('3. Retrasos')
                        print("4. Préstamos por retornar")
                        print("5. Préstamos por período")
                        print("6. Regresar al menú")
                        sub_opcion = input("Seleccione una opción: ")
                        if sub_opcion == "1":
                            while True:
                                print('\nRuta: Menú principal -> Informes -> Reportes -> Clientes ')
                                print('1. Reporte completo de clientes')
                                print('2. Cliente específico')
                                print('3. Regresar al menú')
                                sub_opcion = input('Seleccione una opción: ')
                                if sub_opcion == "1":
                                    reporte_clientes()
                                elif sub_opcion == "2":
                                    reporte_clientes_especifico()
                                elif sub_opcion == "3":
                                    break
                                else:
                                    print("Opción no válida")
                        elif sub_opcion == '2':
                            while True:
                                print('\nRuta: Menú principal -> Informes -> Reportes -> Listado Unidades ')
                                print('\nListado de unidades')
                                print('1. Completo')
                                print('2. Por rodada')
                                print('3. Por color')
                                print('4. Volver al menú anterior')
                                sub_opcion_2 = input('Seleccione una opción: ')
                                if sub_opcion_2 == '1':
                                    try:
                                        reporte_unidades()
                                    except IndexError:
                                        print('No hay registros de unidades.')
                                elif sub_opcion_2 == '2':
                                    reporte_por_rodada()
                                elif sub_opcion_2 == '3':
                                    reporte_color()
                                elif sub_opcion_2 == '4':
                                    break
                                else:
                                    print('Opción no valida')
                        elif sub_opcion == '3':
                            try:
                                print('\nRuta: Menú principal -> Informes -> Reportes -> Reporte de Retrasos ')
                                reporte_retrasos("Negocio_Rentas_Bicicletas.db")
                            except UnboundLocalError:
                                print('No hay prestamos con retrasos')
                        elif sub_opcion == "4":
                            print('\nRuta: Menú principal -> Informes -> Reportes -> Reporte Prestamos Por Retornar ')
                            reporte_prestamos_por_retornar()
                        elif sub_opcion == "5":
                            print('\nRuta: Menú principal -> Informes -> Reportes -> Reporte Prestamos Por Periodo ')
                            reporte_prestamos_por_periodo()
                        elif sub_opcion == "6":
                            break
                        else:
                            print("Opción no válida.")
                elif sub_opcion == '2':
                    while True:
                        print('\nRuta: Menú principal -> Informes -> Análisis')
                        print('\nAnálisis')
                        print('1. Duración de los prestamos')
                        print('2. Ranking de los clientes')
                        print('3. Preferencia de rentas')
                        print('4. Volver al menú de informes')
                        sub_sub_opcion = input('Seleccione una opción: ')
                        if sub_sub_opcion == '1':
                            print('\nRuta: Menú principal -> Informes -> Análisis -> Duración de Análisis')
                            duracion_prestamo()
                        elif sub_sub_opcion == '2':
                            try:
                                print('\nRuta: Menú principal -> Informes -> Análisis -> Ranking Clientes')
                                ranking_clientes()
                            except IndexError:
                                print('No hay clientes registrados') 
                        elif sub_sub_opcion == '3':
                            print('\nRuta: Menú principal -> Informes -> Análisis -> Preferencia de Rentas')
                            while True:
                                print('1. Por rodada')
                                print('2. Por color')
                                print('3. Por día de la semana')
                                print('4. Volver al menú de análisis')
                                sub_opcion_3 = input('Seleccione una opción: ')
                                if sub_opcion_3 == '1':
                                    try:
                                        print('\nRuta: Menú principal -> Informes -> Análisis -> Preferencia de Rentas -> Por rodada ')
                                        reporte_prestamos_por_rodada()
                                    except IndexError:
                                        print('No hay prestamos.')    
                                if sub_opcion_3 == '2':
                                    try:
                                        print('\nRuta: Menú principal -> Informes -> Análisis -> Preferencia de Rentas -> Por color ')
                                        analisis_color()
                                    except IndexError:
                                        print('No hay prestamos registrados.')
                                if sub_opcion_3 == '3':
                                    try:
                                        print('\nRuta: Menú principal -> Informes -> Análisis -> Preferencia de Rentas -> Por día de la semana ')
                                        dia_semana()
                                    except IndexError:
                                        print('No hay prestamos.')
                                elif sub_opcion_3 == '4':
                                    break
                        elif sub_sub_opcion == "4":
                            break
                        else:
                            print("Opción no válida.")
                elif sub_opcion == "3":
                    break
                else:
                    print("Opción no válida.")

        elif opcion == "5":
            confirmacion = input("¿Está seguro de que desea salir? (S/n): ")
            if confirmacion.strip().lower() == 's':
                print("Datos guardados con éxito. Saliendo del sistema.")
                break
            else:
                print("Operación cancelada. Regresando al menú principal.")
        else:
            print("Opción no válida.")

if __name__ == "__main__":
    menu()